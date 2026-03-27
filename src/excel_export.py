"""
excel_export.py -- Excel workbook creation with natively calculated VaR/ES formulas.
"""

from __future__ import annotations

import math
import os
from datetime import date
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from loguru import logger

# Cell values can be str, int, float, None, etc.
CellValue = str | int | float | None
SummaryRow = list[CellValue]


# ---------------------------------------------------------------------------
# Shared styles
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
VAR_95_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # orangish-yellow
VAR_99_FILL = PatternFill(start_color="F6C96C", end_color="F6C96C", fill_type="solid")  # light amber
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Public helpers
# ---------------------------------------------------------------------------


def make_output_dir() -> str:
    """Create and return the directory ``output/{YYYY-MM-DD}/``."""
    dir_path = os.path.join("output", date.today().strftime("%Y-%m-%d"))
    os.makedirs(dir_path, exist_ok=True)
    logger.debug(f"Output directory: {dir_path}")
    return dir_path


# ---------------------------------------------------------------------------
# Data columns (A-D) -- shared between Historical and Parametric
# ---------------------------------------------------------------------------


def _write_data_columns(
    worksheet: Worksheet,
    prices: pd.Series,
    max_data_row: int,
    method: str,
) -> None:
    """Write date/price headers and rows into columns A-D.

    Column C:
    - Historical: -(P_t - P_{t-1}) / P_{t-1}  (arithmetic loss, positive = loss)
    - Parametric: LN(P_t / P_{t-1})            (log return, negative = loss)
    Column D:
    - Historical: LARGE() -- losses descending (worst first)
    - Parametric: SMALL() -- returns ascending (worst first)
    """
    if method == "Historical":
        col_c_header = "Daily Arithmetic Return"
        col_d_header = "Sorted Return"
    else:
        col_c_header = "Daily Log Return"
        col_d_header = "Sorted Return"
    center = Alignment(horizontal="right")

    headers = ["Date", "Close Price", col_c_header, col_d_header]
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        if col_idx in (2, 3, 4):
            cell.alignment = center

    dates = pd.DatetimeIndex(prices.index)
    price_values = prices.values

    for i in range(len(prices)):
        row = i + 2
        worksheet.cell(row=row, column=1, value=dates[i].strftime("%Y-%m-%d"))
        price_cell = worksheet.cell(row=row, column=2, value=float(price_values[i]))
        price_cell.alignment = center

        if row > 2:
            if method == "Historical":
                col_c_formula = f"=(B{row}-B{row - 1})/B{row - 1}"
                col_d_formula = f"=SMALL(C$3:C${max_data_row}, ROW()-2)"
            else:
                col_c_formula = f"=LN(B{row}/B{row - 1})"
                col_d_formula = f"=SMALL(C$3:C${max_data_row}, ROW()-2)"
            return_cell = worksheet.cell(row=row, column=3, value=col_c_formula)
            return_cell.number_format = "0.0000%"
            return_cell.alignment = center

            sorted_cell = worksheet.cell(row=row, column=4, value=col_d_formula)
            sorted_cell.number_format = "0.0000%"
            sorted_cell.alignment = center


# ---------------------------------------------------------------------------
# VaR / ES formulas -- method-specific
# ---------------------------------------------------------------------------


def _var_dollar_formula(method: str, max_data_row: int, alpha: float, pv_ref: str) -> str:
    """Return the Excel formula for 1-Day VaR ($) — positive = loss.

    Historical: PERCENTILE(losses, confidence) * V
    Parametric: -V * (mu - z_alpha * sigma)   where column C = LN returns
    """
    confidence = 1.0 - alpha
    rng = f"C$3:C${max_data_row}"
    if method == "Historical":
        return f"=-PERCENTILE({rng},{alpha})*{pv_ref}"
    else:
        return (
            f"=-{pv_ref}*(AVERAGE({rng})"
            f"-_xlfn.NORM.S.INV({confidence})*_xlfn.STDEV.S({rng}))"
        )


def _es_dollar_formula(method: str, max_data_row: int, alpha: float, pv_ref: str) -> str:
    """Return the Excel formula for 1-Day ES ($) — positive = loss.

    Historical: ES = E[loss | loss > VaR] where loss = -return
                -AVERAGEIF(returns < VaR_threshold) * V
    Parametric: -V * (mu - sigma * phi(z) / alpha)
    """
    confidence = 1.0 - alpha
    rng = f"C$3:C${max_data_row}"
    if method == "Historical":
        var_threshold = f"PERCENTILE({rng},{alpha})"
        return f'=-AVERAGEIF({rng},"<"&{var_threshold})*{pv_ref}'
    else:
        return (
            f"=-{pv_ref}*(AVERAGE({rng})"
            f"-_xlfn.STDEV.S({rng})"
            f"*_xlfn.NORM.DIST(_xlfn.NORM.S.INV({confidence}),0,1,FALSE)/{alpha})"
        )


# ---------------------------------------------------------------------------
# Core export (shared between Historical and Parametric)
# ---------------------------------------------------------------------------


def _export_sheet(
    method: str,
    path: str,
    prices: pd.Series,
    ticker: str,
    n_days: int,
    portfolio_value: float,
    var_date: pd.Timestamp | None,
    stressed: bool,
    lookback: int | None,
    stress_start: str,
    stress_end: str,
    stress_label: str,
    var_confidence: float = 0.99,
    es_confidence: float = 0.975,
) -> str:
    """Create or append a VaR/ES sheet to an Excel workbook.

    ``method`` must be ``"Historical"`` or ``"Parametric"``.
    """

    # ---- 1. Workbook / sheet setup -----------------------------------------

    sheet_title = "VaR and ES"
    stressed_sheet_title = "Stressed VaR and ES"

    if stressed:
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.create_sheet(title=stressed_sheet_title)
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = sheet_title

    max_data_row = len(prices) + 1

    # ---- 2. Write data columns A-D -----------------------------------------

    _write_data_columns(worksheet, prices, max_data_row, method)

    # ---- 2b. Highlight sorted returns at 95% and 99% VaR positions ---------

    n_returns = max_data_row - 2  # returns start at row 3
    for alpha, fill in [(0.05, VAR_95_FILL), (0.01, VAR_99_FILL)]:
        pos = alpha * n_returns
        lo = math.floor(pos)
        hi = math.ceil(pos)
        for k in {lo, hi}:
            if 1 <= k <= n_returns:
                worksheet.cell(row=k + 2, column=4).fill = fill  # column D

    # ---- 3. Build parameter entries ----------------------------------------

    date_str = var_date.strftime("%Y-%m-%d") if var_date is not None else ""

    if stressed:
        param_entries: list[SummaryRow] = [
            ["Method", method, ""],
            ["Ticker", ticker, ""],
            ["VaR Date", date_str, ""],
            ["Portfolio Value ($)", portfolio_value, ""],
            ["N-Day Horizon", n_days, ""],
            ["Stress Period Start Date", stress_start, ""],
            ["Stress Period End Date", stress_end, ""],
            ["Stress Period", stress_label, ""],
        ]
    else:
        param_entries = [
            ["Method", method, ""],
            ["Ticker", ticker, ""],
            ["VaR Date", date_str, ""],
            ["Portfolio Value ($)", portfolio_value, ""],
            ["N-Day Horizon", n_days, ""],
            ["Return Observations", lookback - 1 if lookback else "", ""],
        ]

    # ---- 4. Compute row layout ---------------------------------------------
    #
    # Layout (0-based indices into summary_data):
    #   [0]       Parameter header
    #   [1..N]    param_entries   (N = len(param_entries))
    #   [N+1]     separator
    #   [N+2]     Standard header  (Risk Metric | 99% VaR ($) | 97.5% ES ($))
    #   [N+3]     Standard 1-Day
    #   [N+4]     Standard 10-Day Scaled
    #   [N+5]     separator
    #   [N+6]     Custom header   (Risk Metric | VaR X% | ES Y%)
    #   [N+7]     Custom 1-Day
    #   [N+8]     Custom 10-Day Scaled
    #   [N+9]     Custom n-Day Scaled  (only when n_days != 10)

    summary_start_row = 2
    summary_start_col = 7  # Column G

    N_params = len(param_entries)
    # Portfolio Value is always param_entries[3]; entries start at absolute index 1
    portfolio_value_abs_idx = 1 + 3  # = 4
    pv_ref = f"$H${summary_start_row + portfolio_value_abs_idx}"

    param_header_idx = 0
    param_separator_idx = N_params + 1

    std_header_idx = param_separator_idx + 1
    std_1day_idx = std_header_idx + 1
    std_10day_idx = std_1day_idx + 1
    std_separator_idx = std_10day_idx + 1

    custom_header_idx = std_separator_idx + 1
    custom_1day_idx = custom_header_idx + 1
    custom_10day_idx = custom_1day_idx + 1
    custom_nday_idx: int | None = None
    if n_days != 10:
        custom_nday_idx = custom_10day_idx + 1

    # ---- 5. Build standard table (fixed 99% VaR / 97.5% ES) ----------------

    # Only show custom table when the selected levels differ from the standard (99%/97.5%)
    show_custom = not (var_confidence == 0.99 and es_confidence == 0.975)

    std_1day_h = f"H{summary_start_row + std_1day_idx}"
    std_1day_i = f"I{summary_start_row + std_1day_idx}"

    std_header_label = "Standard Stressed Risk Summary" if stressed else "Standard Risk Summary"

    std_rows: list[SummaryRow] = [
        [std_header_label, "99% VaR ($)", "97.5% ES ($)"],
        ["1-Day",
         _var_dollar_formula(method, max_data_row, 0.01, pv_ref),
         _es_dollar_formula(method, max_data_row, 0.025, pv_ref)],
        ["10-Day Scaled",
         f"={std_1day_h} * SQRT(10)",
         f"={std_1day_i} * SQRT(10)"],
    ]
    std_nday_idx: int | None = None
    if n_days != 10 and not show_custom:
        std_nday_idx = std_10day_idx + 1
        std_rows.append([
            f"{n_days}-Day Scaled",
            f"={std_1day_h} * SQRT({n_days})",
            f"={std_1day_i} * SQRT({n_days})",
        ])

    # ---- 6. Build custom table (user-selected confidence levels) ------------

    var_alpha = 1.0 - var_confidence
    es_alpha = 1.0 - es_confidence
    var_conf_label = f"{var_confidence * 100:g}%"
    es_conf_label = f"{es_confidence * 100:g}%"

    custom_1day_h = f"H{summary_start_row + custom_1day_idx}"
    custom_1day_i = f"I{summary_start_row + custom_1day_idx}"

    custom_header_label = "Stressed Risk Summary" if stressed else "Risk Summary"

    custom_rows: list[SummaryRow] = [
        [custom_header_label, f"{var_conf_label} VaR ($)", f"{es_conf_label} ES ($)"],
        ["1-Day",
         _var_dollar_formula(method, max_data_row, var_alpha, pv_ref),
         _es_dollar_formula(method, max_data_row, es_alpha, pv_ref)],
        ["10-Day Scaled",
         f"={custom_1day_h} * SQRT(10)",
         f"={custom_1day_i} * SQRT(10)"],
    ]
    if n_days != 10:
        custom_rows.append([
            f"{n_days}-Day Scaled",
            f"={custom_1day_h} * SQRT({n_days})",
            f"={custom_1day_i} * SQRT({n_days})",
        ])

    # ---- 7. Assemble full summary ------------------------------------------

    empty_row: SummaryRow = ["", "", ""]
    summary_data: list[SummaryRow] = []
    summary_data.append(["Parameter", "Value", ""])       # index 0
    summary_data.extend(param_entries)                    # indices 1..N
    summary_data.append(empty_row)                        # index N+1
    summary_data.extend(std_rows)                         # indices N+2..N+4
    if show_custom:
        summary_data.append(empty_row)                    # index N+5
        summary_data.extend(custom_rows)                  # indices N+6..

    # ---- 9. Write and style the summary ------------------------------------

    # Indices of param-section rows (skip col I for these)
    param_all_indices = {param_header_idx} | set(range(1, 1 + N_params))

    # Which rows to right-align value (col H) — param entries except Portfolio Value
    right_align_indices = set(range(1, 1 + N_params)) - {portfolio_value_abs_idx}

    # Which rows get money ($) formatting
    money_format_indices: set[int] = {
        portfolio_value_abs_idx,
        std_1day_idx, std_10day_idx,
    }
    if std_nday_idx is not None:
        money_format_indices.add(std_nday_idx)
    if show_custom:
        money_format_indices |= {custom_1day_idx, custom_10day_idx}
        if custom_nday_idx is not None:
            money_format_indices.add(custom_nday_idx)

    # Which rows get dark-blue header styling
    section_header_indices = {param_header_idx, std_header_idx}
    if show_custom:
        section_header_indices.add(custom_header_idx)

    # Which rows skip all styling (separators)
    unstyled_indices = {param_separator_idx}
    if show_custom:
        unstyled_indices.add(std_separator_idx)

    for data_index, row_data in enumerate(summary_data):
        sheet_row = summary_start_row + data_index
        for col_offset, value in enumerate(row_data):
            # Third column (col I) is unused for parameter rows
            if data_index in param_all_indices and col_offset == 2:
                continue

            col = summary_start_col + col_offset
            cell = worksheet.cell(row=sheet_row, column=col, value=value)  # type: ignore[arg-type]

            # Skip styling for empty separators
            if data_index in unstyled_indices:
                continue

            cell.border = THIN_BORDER

            if data_index in section_header_indices:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                if col == 8:
                    cell.alignment = Alignment(horizontal="right")
                elif col == 9 and data_index != param_header_idx:
                    cell.alignment = Alignment(horizontal="right")

            if data_index in right_align_indices and col_offset == 1:
                cell.alignment = Alignment(horizontal="right")

            if col in (8, 9):  # Columns H and I
                if data_index in money_format_indices:
                    cell.number_format = '"$"#,##0.00'

    # ---- 10. Column widths -------------------------------------------------

    column_widths = {
        "A": 12, "B": 15, "C": 20, "D": 15,
        "E": 5, "F": 5,
        "G": 27, "H": 24, "I": 24,
    }
    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    # ---- 11. Save ----------------------------------------------------------

    workbook.save(path)
    action = "sheet added" if stressed else "report saved"
    logger.info(f"{method} VaR ES {action}: {path}")
    return path


# ---------------------------------------------------------------------------
# API -- thin wrappers around _export_sheet
# ---------------------------------------------------------------------------


def export_historical_var_sheet(
    path: str,
    prices: pd.Series,
    ticker: str,
    n_days: int,
    portfolio_value: float,
    var_date: pd.Timestamp | None = None,
    stressed: bool = False,
    lookback: int | None = None,
    stress_start: str = "",
    stress_end: str = "",
    stress_label: str = "",
    var_confidence: float = 0.99,
    es_confidence: float = 0.975,
) -> str:
    """Create or append a Historical VaR/ES sheet to an Excel workbook."""
    return _export_sheet(
        "Historical", path, prices, ticker, n_days, portfolio_value,
        var_date, stressed, lookback, stress_start, stress_end, stress_label,
        var_confidence, es_confidence,
    )


def export_parametric_var_sheet(
    path: str,
    prices: pd.Series,
    ticker: str,
    n_days: int,
    portfolio_value: float,
    var_date: pd.Timestamp | None = None,
    stressed: bool = False,
    lookback: int | None = None,
    stress_start: str = "",
    stress_end: str = "",
    stress_label: str = "",
    var_confidence: float = 0.99,
    es_confidence: float = 0.975,
) -> str:
    """Create or append a Parametric VaR/ES sheet to an Excel workbook."""
    return _export_sheet(
        "Parametric", path, prices, ticker, n_days, portfolio_value,
        var_date, stressed, lookback, stress_start, stress_end, stress_label,
        var_confidence, es_confidence,
    )


# ---------------------------------------------------------------------------
# Report-level exports (output dir + both normal & stressed sheets)
# ---------------------------------------------------------------------------


def export_historical_var_report(
    prices: pd.Series,
    ticker: str,
    n_days: int,
    portfolio_value: float,
    var_date: pd.Timestamp | None,
    lookback: int,
    stressed_prices: pd.Series,
    stress_start: str,
    stress_end: str,
    stress_label: str,
    var_confidence: float = 0.99,
    es_confidence: float = 0.975,
) -> str:
    """Generate a full Historical VaR Excel report (normal + stressed sheets)."""
    output_dir = make_output_dir()
    date_str = var_date.strftime("%Y-%m-%d") if var_date else ""
    excel_path = os.path.join(output_dir, f"{ticker}_{date_str}_Historical_VaR.xlsx")

    export_historical_var_sheet(
        path=excel_path, prices=prices, ticker=ticker, n_days=n_days,
        portfolio_value=portfolio_value, var_date=var_date, stressed=False,
        lookback=lookback, var_confidence=var_confidence, es_confidence=es_confidence,
    )
    export_historical_var_sheet(
        path=excel_path, prices=stressed_prices, ticker=ticker, n_days=n_days,
        portfolio_value=portfolio_value, var_date=var_date, stressed=True,
        stress_start=stress_start, stress_end=stress_end, stress_label=stress_label,
        var_confidence=var_confidence, es_confidence=es_confidence,
    )
    return excel_path


def export_parametric_var_report(
    prices: pd.Series,
    ticker: str,
    n_days: int,
    portfolio_value: float,
    var_date: pd.Timestamp | None,
    lookback: int,
    stressed_prices: pd.Series,
    stress_start: str,
    stress_end: str,
    stress_label: str,
    var_confidence: float = 0.99,
    es_confidence: float = 0.975,
) -> str:
    """Generate a full Parametric VaR Excel report (normal + stressed sheets)."""
    output_dir = make_output_dir()
    date_str = var_date.strftime("%Y-%m-%d") if var_date else ""
    excel_path = os.path.join(output_dir, f"{ticker}_{date_str}_Parametric_VaR.xlsx")

    export_parametric_var_sheet(
        path=excel_path, prices=prices, ticker=ticker, n_days=n_days,
        portfolio_value=portfolio_value, var_date=var_date, stressed=False,
        lookback=lookback, var_confidence=var_confidence, es_confidence=es_confidence,
    )
    export_parametric_var_sheet(
        path=excel_path, prices=stressed_prices, ticker=ticker, n_days=n_days,
        portfolio_value=portfolio_value, var_date=var_date, stressed=True,
        stress_start=stress_start, stress_end=stress_end, stress_label=stress_label,
        var_confidence=var_confidence, es_confidence=es_confidence,
    )
    return excel_path
